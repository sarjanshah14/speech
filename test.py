#!/usr/bin/env python3
"""
Refactored Live Google Speech-to-Text Transcriber
Layered Architecture with Universal Error Handling and Recovery
"""

import os
import sys
import time
import tempfile
import wave
import re
import argparse
from typing import Tuple, List, Dict, Optional, Callable
from difflib import SequenceMatcher
from enum import Enum
from functools import wraps

import numpy as np
import sounddevice as sd
import openpyxl
from google.cloud import speech

# Set Google credentials
os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = r"C:\Users\itintern2\Desktop\project1\keys\vigilant-art-475111-c1-cd0522c17ae2.json"

# AUDIO PARAMETERS
SAMPLE_RATE = 16000
CHANNELS = 1
DTYPE = "int16"
SILENCE_SECONDS = 1.5
SILENCE_AMPLITUDE = 300

# Global debug flag
DEBUG = False


class ProcessingStage(Enum):
    """Enumeration for processing stages"""
    RAW_CLEANUP = "Stage 1: Raw Cleanup"
    LINGUISTIC_NORMALIZATION = "Stage 2: Linguistic Normalization"
    PATTERN_EXTRACTION = "Stage 3: Pattern Extraction"
    CONFIDENCE_RECOVERY = "Stage 4: Confidence & Recovery"


class ErrorHandler:
    """Unified error handling and logging"""
    
    @staticmethod
    def log_error(stage: ProcessingStage, error: Exception, context: str = ""):
        """Log error with context"""
        print(f"\n❌ ERROR in {stage.value}")
        print(f"   Error: {type(error).__name__}: {error}")
        if context:
            print(f"   Context: {context}")
        if DEBUG:
            import traceback
            traceback.print_exc()
    
    @staticmethod
    def safe_execute(func: Callable, stage: ProcessingStage, default_return=None, context: str = ""):
        """Execute function with error handling"""
        try:
            return func()
        except Exception as e:
            ErrorHandler.log_error(stage, e, context)
            return default_return
    
    @staticmethod
    def error_handler(stage: ProcessingStage):
        """Decorator for error handling"""
        def decorator(func):
            @wraps(func)
            def wrapper(*args, **kwargs):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    ErrorHandler.log_error(stage, e, f"Function: {func.__name__}")
                    return None
            return wrapper
        return decorator


class RawCleanupLayer:
    """Stage 1: Basic text sanitation and cleanup"""
    
    def __init__(self, debug: bool = False):
        self.debug = debug
        # Common product code prefixes for better handling
        self.product_prefixes = {
            'in': 'IN', 'ip': 'IP', 'ai': 'AI', 'blp': 'BLP', 
            'jc': 'JC', 's1': 'S1', 'c21': 'C21'
        }
    
    def process(self, text: str) -> str:
        """Stage 1: Raw cleanup - basic text sanitation"""
        if self.debug:
            print(f"  [{ProcessingStage.RAW_CLEANUP.value}] Input: '{text}'")
        
        # Convert to lowercase and strip
        text = text.lower().strip()
        
        # Remove unwanted punctuation (keep dots and spaces)
        text = re.sub(r'[,!?;:\'"()]', '', text)
        
        # Remove extra whitespace
        text = ' '.join(text.split())
        
        # Common transcription noise cleanup
        text = re.sub(r'\bp\s+c\b', 'pieces', text)
        text = re.sub(r'\bpc\b', 'pieces', text)
        text = re.sub(r'\bzed\b', 'z', text)
        text = re.sub(r'\band\b', ' ', text)
        
        # Fix common misheard words (only for non-product-code contexts)
        text = re.sub(r'\bcome\b', 'cum', text)
        text = re.sub(r'\bpace\b', 'piece', text)
        text = re.sub(r'\bpage\b', 'piece', text)
        
        # ENHANCED: Product code specific fixes
        # Handle common product code prefixes that might be misheard
        for spoken, written in self.product_prefixes.items():
            # Only replace if it's at the start of a word/phrase
            text = re.sub(rf'\b{spoken}\b', written, text)
        
        # Handle "double zero" → "00" for product codes
        text = re.sub(r'\bdouble\s+zero\b', '00', text)
        text = re.sub(r'\bdouble\s+oh\b', '00', text)
        
        # Handle "triple zero" → "000" for product codes
        text = re.sub(r'\btriple\s+zero\b', '000', text)
        text = re.sub(r'\btriple\s+oh\b', '000', text)
        
        if self.debug:
            print(f"  [{ProcessingStage.RAW_CLEANUP.value}] Output: '{text}'")
        
        return text


class LinguisticNormalizer:
    """Stage 2: Linguistic normalization with number and pattern conversion"""
    
    def __init__(self, debug: bool = False):
        self.debug = debug
        self.number_words = {
            'zero': '0', 'oh': '0',
            'one': '1', 'two': '2', 'three': '3', 'four': '4',
            'five': '5', 'six': '6', 'seven': '7', 'eight': '8', 'nine': '9',
            'ten': '10', 'eleven': '11', 'twelve': '12', 'thirteen': '13',
            'fourteen': '14', 'fifteen': '15', 'sixteen': '16',
            'seventeen': '17', 'eighteen': '18', 'nineteen': '19',
            'twenty': '20', 'thirty': '30', 'forty': '40', 'fifty': '50',
            'sixty': '60', 'seventy': '70', 'eighty': '80', 'ninety': '90'
        }
        
        # Extended number mappings
        self.word_to_digit = {
            **self.number_words,
            'to': '2', 'too': '2', 'for': '4', 'ate': '8'
        }
    
    def process(self, text: str) -> str:
        """Stage 2: Linguistic normalization"""
        if self.debug:
            print(f"  [{ProcessingStage.LINGUISTIC_NORMALIZATION.value}] Input: '{text}'")
        
        original_text = text
        
        # FIX 1: Handle "dot" patterns BEFORE number conversions to preserve structure
        # "s dot ss dot de" → "s.ss.de"
        text = re.sub(r'\bdots?\s+dots?\b', '..', text)
        text = re.sub(r'\bdots?\b', '.', text)
        
        # ENHANCED: Handle spaced letters that should be combined
        # "i n b a c a 0 1" → "inbaca01"
        # Simple approach: combine consecutive single letters
        while True:
            # Find pattern: single letter, space, single letter
            match = re.search(r'\b([a-z])\s+([a-z])\b', text)
            if not match:
                break
            # Replace with combined letters
            text = text.replace(match.group(0), match.group(1) + match.group(2))
        
        # Convert common prefixes to uppercase
        text = re.sub(r'\bin\b', 'IN', text)
        text = re.sub(r'\bip\b', 'IP', text)
        
        # Continue combining remaining spaced letters
        while True:
            # Find pattern: letters, space, letters
            match = re.search(r'\b([a-z]+)\s+([a-z]+)\b', text)
            if not match:
                break
            # Replace with combined letters
            text = text.replace(match.group(0), match.group(1) + match.group(2))
        
        # Handle spaced numbers: "0 1" → "01"
        while True:
            # Find pattern: digit, space, digit
            match = re.search(r'\b(\d)\s+(\d)\b', text)
            if not match:
                break
            # Replace with combined digits
            text = text.replace(match.group(0), match.group(1) + match.group(2))
        
        # Final cleanup: remove remaining spaces between letters and numbers
        text = re.sub(r'\b([A-Za-z]+)\s+(\d+)\b', r'\1\2', text)
        
        # Remove any remaining spaces
        text = text.replace(' ', '')
        
        if self.debug:
            print(f"  [NORM] After dot handling: '{text}'")
        
        # FIX 2: Handle double/triple BEFORE simple number conversion to avoid conflicts
        # This must happen before number word conversion
        text = self._handle_double_triple(text)
        
        if self.debug:
            print(f"  [NORM] After double/triple: '{text}'")
        
        # Handle compound numbers (twenty one, thirty five, etc.)
        text = self._handle_compound_numbers(text)
        
        if self.debug:
            print(f"  [NORM] After compound numbers: '{text}'")
        
        # FIX 3: Convert number words to digits
        # BUT: Prevent accidental conversion of quantity patterns like "three three" → "33"
        # We need to be careful not to convert when it would cause math errors
        text = self._convert_number_words_safe(text)
        
        if self.debug:
            print(f"  [NORM] After number word conversion: '{text}'")
        
        # FIX 4: Handle "n" vs "in" - DO NOT convert "in" to "n"
        # If user says "in", keep it as "in". Only if user explicitly says "n", it's "n"
        # No automatic conversion!
        
        # Special case: "eight" at start might be "H" (only if followed by number word)
        text = self._fix_eight_as_h(text)
        
        if self.debug:
            print(f"  [NORM] After special fixes: '{text}'")
        
        # Keep only letters, digits, dots, and spaces
        text = re.sub(r'[^\w\s.]', '', text)
        
        # Final cleanup: normalize spaces around dots
        text = re.sub(r'\s*\.\s*', '.', text)
        text = re.sub(r'\s+', ' ', text).strip()
        
        if self.debug:
            print(f"  [{ProcessingStage.LINGUISTIC_NORMALIZATION.value}] Output: '{text}'")
        
        return text
    
    def _handle_double_triple(self, text: str) -> str:
        """Handle double/triple patterns correctly"""
        # Check if "double" or "triple" exists in text
        if 'double' not in text and 'triple' not in text:
            return text
        
        # FIX: Handle "double 2 3" → "223"
        # Pattern: double number number → doubled number + number
        text = re.sub(r'\bdouble\s+(\d)\s+(\d)\b', r'\1\1\2', text)
        
        # Pattern: number(s) double number → numbers + doubled number
        text = re.sub(r'(\d+)\s+double\s+(\d)', r'\1\2\2', text)
        
        # Handle "triple 0 3" → "0003" (not "03"!)
        # Pattern: letter(s) triple number number → letter(s) + tripled number + number
        # Example: "z triple 0 3" → "z" + "000" + "3" = "z0003"
        text = re.sub(r'([a-z]+)\s+triple\s+(\d)\s+(\d)', r'\1\2\2\2\3', text)
        # Also handle without letter prefix: "triple 0 3" → "0003"
        text = re.sub(r'(?:^|\s)triple\s+(\d)\s+(\d)', r'\1\1\1\2', text)
        
        # Handle "double digit" and "triple digit" patterns directly
        text = re.sub(r'\bdouble\s+(\d)\b', r'\1\1', text)
        text = re.sub(r'\btriple\s+(\d)\b', r'\1\1\1', text)
        
        # Handle "X double Y" → "XYY" (where X and Y are number words)
        for num1_word, num1_digit in self.number_words.items():
            for num2_word, num2_digit in self.number_words.items():
                # Pattern: word1 double word2
                pattern = rf'\b{re.escape(num1_word)}\s+double\s+{re.escape(num2_word)}\b'
                replacement = num1_digit + num2_digit + num2_digit
                text = re.sub(pattern, replacement, text)
                
                # Pattern: letter(s) word1 double word2
                pattern = rf'\b([a-z]+)\s+{re.escape(num1_word)}\s+double\s+{re.escape(num2_word)}\b'
                def repl_func(match):
                    return match.group(1) + num1_digit + num2_digit + num2_digit
                text = re.sub(pattern, repl_func, text)
        
        # Handle "X triple Y" → "XYYY" (where X and Y are number words)
        for num1_word, num1_digit in self.number_words.items():
            for num2_word, num2_digit in self.number_words.items():
                # Pattern: word1 triple word2
                pattern = rf'\b{re.escape(num1_word)}\s+triple\s+{re.escape(num2_word)}\b'
                replacement = num1_digit + num2_digit + num2_digit + num2_digit
                text = re.sub(pattern, replacement, text)
                
                # Pattern: letter(s) word1 triple word2
                pattern = rf'\b([a-z]+)\s+{re.escape(num1_word)}\s+triple\s+{re.escape(num2_word)}\b'
                def repl_func(match):
                    return match.group(1) + num1_digit + num2_digit + num2_digit + num2_digit
                text = re.sub(pattern, repl_func, text)
        
        # Handle standalone "double X" → "XX"
        for word, digit in self.number_words.items():
            pattern = rf'\bdouble\s+{re.escape(word)}\b'
            text = re.sub(pattern, digit + digit, text)
        
        # Handle standalone "triple X" → "XXX"
        for word, digit in self.number_words.items():
            pattern = rf'\btriple\s+{re.escape(word)}\b'
            text = re.sub(pattern, digit + digit + digit, text)
        
        # Handle "double digit" and "triple digit" patterns directly
        text = re.sub(r'\bdouble\s+(\d)\b', r'\1\1', text)
        text = re.sub(r'\btriple\s+(\d)\b', r'\1\1\1', text)
        
        return text
    
    def _handle_compound_numbers(self, text: str) -> str:
        """Handle compound numbers like twenty-one, thirty-five"""
        compound_patterns = [
            (r'\btwenty\s+one\b', '21'), (r'\btwenty\s+two\b', '22'), (r'\btwenty\s+three\b', '23'),
            (r'\btwenty\s+four\b', '24'), (r'\btwenty\s+five\b', '25'), (r'\btwenty\s+six\b', '26'),
            (r'\btwenty\s+seven\b', '27'), (r'\btwenty\s+eight\b', '28'), (r'\btwenty\s+nine\b', '29'),
            (r'\bthirty\s+one\b', '31'), (r'\bthirty\s+two\b', '32'), (r'\bthirty\s+three\b', '33'),
            (r'\bthirty\s+four\b', '34'), (r'\bthirty\s+five\b', '35'), (r'\bthirty\s+six\b', '36'),
            (r'\bthirty\s+seven\b', '37'), (r'\bthirty\s+eight\b', '38'), (r'\bthirty\s+nine\b', '39'),
            (r'\bforty\s+one\b', '41'), (r'\bforty\s+two\b', '42'), (r'\bforty\s+three\b', '43'),
            (r'\bforty\s+four\b', '44'), (r'\bforty\s+five\b', '45'), (r'\bforty\s+six\b', '46'),
            (r'\bforty\s+seven\b', '47'), (r'\bforty\s+eight\b', '48'), (r'\bforty\s+nine\b', '49'),
            (r'\bfifty\s+one\b', '51'), (r'\bfifty\s+two\b', '52'), (r'\bfifty\s+three\b', '53'),
            (r'\bfifty\s+four\b', '54'), (r'\bfifty\s+five\b', '55'), (r'\bfifty\s+six\b', '56'),
            (r'\bfifty\s+seven\b', '57'), (r'\bfifty\s+eight\b', '58'), (r'\bfifty\s+nine\b', '59'),
            (r'\bsixty\s+one\b', '61'), (r'\bsixty\s+two\b', '62'), (r'\bsixty\s+three\b', '63'),
            (r'\bsixty\s+four\b', '64'), (r'\bsixty\s+five\b', '65'), (r'\bsixty\s+six\b', '66'),
            (r'\bsixty\s+seven\b', '67'), (r'\bsixty\s+eight\b', '68'), (r'\bsixty\s+nine\b', '69'),
            (r'\bseventy\s+one\b', '71'), (r'\bseventy\s+two\b', '72'), (r'\bseventy\s+three\b', '73'),
            (r'\bseventy\s+four\b', '74'), (r'\bseventy\s+five\b', '75'), (r'\bseventy\s+six\b', '76'),
            (r'\bseventy\s+seven\b', '77'), (r'\bseventy\s+eight\b', '78'), (r'\bseventy\s+nine\b', '79'),
            (r'\beighty\s+one\b', '81'), (r'\beighty\s+two\b', '82'), (r'\beighty\s+three\b', '83'),
            (r'\beighty\s+four\b', '84'), (r'\beighty\s+five\b', '85'), (r'\beighty\s+six\b', '86'),
            (r'\beighty\s+seven\b', '87'), (r'\beighty\s+eight\b', '88'), (r'\beighty\s+nine\b', '89'),
            (r'\bninety\s+one\b', '91'), (r'\bninety\s+two\b', '92'), (r'\bninety\s+three\b', '93'),
            (r'\bninety\s+four\b', '94'), (r'\bninety\s+five\b', '95'), (r'\bninety\s+six\b', '96'),
            (r'\bninety\s+seven\b', '97'), (r'\bninety\s+eight\b', '98'), (r'\bninety\s+nine\b', '99'),
        ]
        
        for pattern, replacement in compound_patterns:
            text = re.sub(pattern, replacement, text)
        
        return text
    
    def _convert_number_words_safe(self, text: str) -> str:
        """Convert number words to digits, but prevent accidental quantity errors"""
        # FIX: Prevent "3 three" or "three three" from becoming "33" or being interpreted as math
        # These patterns should be preserved as-is or handled differently
        
        # First, mark patterns that should NOT be converted
        # Pattern: digit followed by same number word (e.g., "3 three" or "three three")
        # These are likely quantity errors and should not be auto-corrected
        protected_patterns = []
        for word, digit in self.word_to_digit.items():
            # Protect "digit word" and "word word" patterns
            pattern1 = rf'(\d)\s+{word}\b'
            pattern2 = rf'\b{word}\s+{word}\b'
            
            # Store original matches
            for match in re.finditer(pattern1, text):
                protected_patterns.append((match.group(), match.start(), match.end()))
            for match in re.finditer(pattern2, text):
                protected_patterns.append((match.group(), match.start(), match.end()))
        
        # Sort by position (reverse) to preserve indices during replacement
        protected_patterns.sort(key=lambda x: x[1], reverse=True)
        
        # Temporarily replace protected patterns with placeholders
        placeholders = {}
        for i, (match_text, start, end) in enumerate(protected_patterns):
            placeholder = f"__PROTECTED_{i}__"
            placeholders[placeholder] = match_text
            text = text[:start] + placeholder + text[end:]
        
        # Now convert remaining number words
        for word, digit in sorted(self.word_to_digit.items(), key=lambda x: len(x[0]), reverse=True):
            text = re.sub(rf'\b{word}\b', digit, text)
        
        # Restore protected patterns (they stay as-is to avoid auto-correction)
        for placeholder, original in placeholders.items():
            text = text.replace(placeholder, original)
        
        return text
    
    def _fix_eight_as_h(self, text: str) -> str:
        """Fix 'eight' at start that might be 'H'"""
        words = text.split()
        if words and words[0] == 'eight' and len(words) > 1:
            digit_words = set(self.number_words.keys())
            if words[1] in digit_words:
                text = 'h ' + ' '.join(words[1:])
                if self.debug:
                    print(f"  [NORM] Corrected 'eight' → 'h': '{text}'")
        return text


class PatternExtractor:
    """Stage 3: Extract product code and quantity patterns"""
    
    def __init__(self, debug: bool = False, allowed_lengths: Optional[set] = None):
        self.debug = debug
        # Set of allowed alphanumeric lengths derived from product list
        self.allowed_lengths = allowed_lengths or set()
    
    def extract_candidates(self, normalized: str, original: str) -> List[str]:
        """Extract all possible product code sequences"""
        if self.debug:
            print(f"  [{ProcessingStage.PATTERN_EXTRACTION.value}] Input: '{normalized}'")
        
        candidates = []
        
        # Remove extra spaces
        clean_text = ' '.join(normalized.split())
        no_space = clean_text.replace(' ', '')
        
        # FIX: Extract spaced numbers correctly
        # "35 335" → "35335", "in c 35 335" → "c35335"
        spaced_numbers = re.findall(r'\d+(?:\s+\d+)+', clean_text)
        for match in spaced_numbers:
            compact = match.replace(' ', '')
            if len(compact) >= 4:
                candidates.append(compact)
                if self.debug:
                    print(f"  [EXTRACT] Spaced number candidate: '{match}' → '{compact}'")
        
        # ENHANCED: Pattern extraction based on actual product code analysis
        # Pattern 1: Codes with dots (e.g., JC.NPP, S1.G.A&T., JC.TM.NPP)
        dot_patterns = re.findall(r'[A-Za-z]+\d*[A-Za-z]*\.[A-Za-z0-9.&]+', no_space)
        candidates.extend(dot_patterns)
        
        # Pattern 2: Single Letter + Numbers (44% of codes: A0002, D0055, H0282)
        single_letter_patterns = re.findall(r'\b[A-Z]\d{4,5}\b', no_space)
        candidates.extend(single_letter_patterns)
        
        # Pattern 3: Multiple Letters + Numbers (17.9% of codes: INBACA01, IP00001, BLP001)
        multi_letter_patterns = re.findall(r'\b[A-Za-z]{2,}\d{2,}\b', no_space)
        candidates.extend(multi_letter_patterns)
        
        # Pattern 4: Pure numeric codes (35.7% of codes: 10006, 12001, 35001)
        # Most are 5 digits, some are longer
        pure_numeric = re.findall(r'\b\d{5,6}\b', no_space)
        candidates.extend(pure_numeric)
        
        # Pattern 5: Mixed patterns (Letters + Numbers + Letters: INCDM1LK)
        mixed_patterns = re.findall(r'\b[A-Z]+\d+[A-Z]+\d*\b', no_space)
        candidates.extend(mixed_patterns)
        
        # Pattern 6: Special patterns like 13ED0001
        special_patterns = re.findall(r'\b\d{2}[A-Z]{2}\d{4}\b', no_space)
        candidates.extend(special_patterns)
        
        # Data-driven candidate generation by allowed lengths (from product list)
        # Any contiguous token of letters/digits/dots whose alnum-length matches
        # known product code lengths becomes a candidate.
        tokens = re.findall(r'[A-Za-z0-9.]+', clean_text)
        for tok in tokens:
            alnum = re.sub(r'[^A-Za-z0-9]', '', tok)
            if self.allowed_lengths and len(alnum) in self.allowed_lengths:
                candidates.append(tok)

        # Remove duplicates while preserving order
        seen = set()
        unique_candidates = []
        for cand in candidates:
            if cand not in seen:
                seen.add(cand)
                unique_candidates.append(cand)
        
        if self.debug:
            print(f"  [{ProcessingStage.PATTERN_EXTRACTION.value}] Candidates: {unique_candidates}")
        
        return unique_candidates


class Matcher:
    """Stage 4: Fuzzy matching and recovery with confidence scoring"""
    
    def __init__(self, pcode_list: List[str], debug: bool = False):
        self.pcode_list = pcode_list
        self.pcode_list_no_dots = [str(p).replace('.', '').lower() for p in pcode_list]
        self.debug = debug
        self.similar_chars = {
            'a': ['e', 'o'], 'b': ['d', 'p', 'v'], 'c': ['s', 'k', 'g'],
            'd': ['b', 't'], 'e': ['a', 'i'], 'f': ['v', 'p', 's'],
            'g': ['j', 'k', 'c'], 'h': ['n', '8'], 'i': ['e', 'y', '1'],
            'j': ['g'], 'k': ['c', 'g', 'q'], 'l': ['r', '1'],
            'm': ['n'], 'n': ['m', 'h'], 'o': ['a', 'u', '0'],
            'p': ['b', 'f'], 'q': ['k'], 'r': ['l'], 's': ['f', 'c', 'z', '5'],
            't': ['d'], 'u': ['o'], 'v': ['f', 'w', 'b'], 'w': ['v', 'u'],
            'x': ['z'], 'y': ['i'], 'z': ['s', 'x'], '0': ['o'],
            '1': ['i', 'l'], '5': ['s'],
        }
        # Cache for recent successful matches
        self.match_cache: Dict[str, str] = {}
    
    def find_best_match(self, candidates: List[str]) -> Tuple[str, str, float]:
        """Find best matching product code with confidence score"""
        if self.debug:
            print(f"  [{ProcessingStage.CONFIDENCE_RECOVERY.value}] Matching {len(candidates)} candidates")
        
        best_match = ""
        best_length = 0
        remainder = ""
        confidence = 0.0
        
        # ENHANCED: Prioritize candidates by pattern type (based on frequency analysis)
        # Sort candidates by pattern priority: Single Letter > Pure Numeric > Multi Letter > Others
        def get_pattern_priority(candidate):
            if re.match(r'^[A-Z]\d{4,5}$', candidate):
                return 1  # Single Letter + Numbers (44% of codes)
            elif re.match(r'^\d{5,6}$', candidate):
                return 2  # Pure Numeric (35.7% of codes)
            elif re.match(r'^[A-Z]{2,}\d{3,}$', candidate):
                return 3  # Multiple Letters + Numbers (17.9% of codes)
            elif '.' in candidate:
                return 4  # With Dots (0.9% of codes)
            else:
                return 5  # Other patterns
        
        # Sort candidates by priority
        candidates.sort(key=get_pattern_priority)
        
        # Try exact matches first
        for candidate in candidates:
            candidate_lower = candidate.lower()
            
            # Check cache first
            if candidate_lower in self.match_cache:
                cached_match = self.match_cache[candidate_lower]
                if self.debug:
                    print(f"  [MATCH] Cache hit: '{candidate}' → '{cached_match}'")
                return cached_match, "", 1.0
            
            # Exact match
            for pcode in self.pcode_list:
                pcode_clean = str(pcode).lower()
                
                if candidate_lower == pcode_clean:
                    if len(pcode_clean) >= best_length:
                        best_length = len(pcode_clean)
                        best_match = str(pcode)
                        remainder = ""
                        confidence = 1.0
                        self.match_cache[candidate_lower] = best_match
                        if self.debug:
                            print(f"  [MATCH] Exact: '{candidate}' → '{best_match}'")
                
                # Prefix match
                elif candidate_lower.startswith(pcode_clean):
                    if len(pcode_clean) > best_length:
                        best_length = len(pcode_clean)
                        best_match = str(pcode)
                        remainder = candidate_lower[len(pcode_clean):]
                        confidence = 0.9
                        if self.debug:
                            print(f"  [MATCH] Prefix: '{candidate}' → '{best_match}' (remainder: {remainder})")
        
        # Try fuzzy matching without dots
        if not best_match:
            for candidate in candidates:
                fuzzy_pcode, fuzzy_remainder = self._fuzzy_match_with_dots(candidate)
                if fuzzy_pcode:
                    fuzzy_length = len(fuzzy_pcode.replace('.', ''))
                    if fuzzy_length > best_length:
                        best_match = fuzzy_pcode
                        remainder = fuzzy_remainder
                        best_length = fuzzy_length
                        confidence = 0.8
                        if self.debug:
                            print(f"  [MATCH] Fuzzy dot: '{candidate}' → '{best_match}'")
        
        # Try character confusion correction
        if not best_match:
            for candidate in candidates:
                similar_matches = self._find_similar_pcodes(candidate)
                if similar_matches:
                    best_match = similar_matches[0]
                    remainder = ""
                    confidence = 0.7
                    if self.debug:
                        print(f"  [MATCH] Similar char: '{candidate}' → '{best_match}'")
                    break
        
        # Recovery mode: Try relaxed patterns
        if not best_match:
            best_match, remainder, confidence = self._recovery_mode(candidates)
        
        return best_match, remainder, confidence
    
    def _fuzzy_match_with_dots(self, candidate: str) -> Tuple[str, str]:
        """Try to match candidate against pcodes with dots removed"""
        candidate_clean = candidate.replace('.', '').replace(' ', '').lower()
        
        for i, pcode_no_dot in enumerate(self.pcode_list_no_dots):
            if candidate_clean == pcode_no_dot:
                return str(self.pcode_list[i]), ""
            
            if candidate_clean.startswith(pcode_no_dot) and len(pcode_no_dot) >= 3:
                remainder = candidate_clean[len(pcode_no_dot):]
                return str(self.pcode_list[i]), remainder
        
        return "", ""
    
    def _find_similar_pcodes(self, pcode: str) -> List[str]:
        """Find similar product codes by replacing confused characters"""
        if not pcode:
            return []
        
        candidates = [pcode.lower()]
        
        # Generate variants
        for i, char in enumerate(pcode.lower()):
            if char in self.similar_chars:
                for similar in self.similar_chars[char]:
                    variant = pcode[:i].lower() + similar + pcode[i+1:].lower()
                    candidates.append(variant)
        
        found = []
        for candidate in candidates:
            candidate_no_dot = candidate.replace('.', '')
            
            for i, pcode_in_list in enumerate(self.pcode_list):
                pcode_lower = str(pcode_in_list).lower()
                pcode_no_dot = self.pcode_list_no_dots[i]
                
                if pcode_lower == candidate or pcode_no_dot == candidate_no_dot:
                    found.append(str(pcode_in_list))
                    break
        
        return found
    
    def _recovery_mode(self, candidates: List[str]) -> Tuple[str, str, float]:
        """Recovery mode: try relaxed patterns and Levenshtein distance"""
        if not candidates:
            return "", "", 0.0
        
        # Try partial matches (at least 3 characters)
        for candidate in candidates:
            candidate_clean = candidate.replace('.', '').lower()
            if len(candidate_clean) < 3:
                continue
            
            for i, pcode_no_dot in enumerate(self.pcode_list_no_dots):
                # Check if candidate is substring or pcode is substring
                if candidate_clean in pcode_no_dot or pcode_no_dot in candidate_clean:
                    similarity = SequenceMatcher(None, candidate_clean, pcode_no_dot).ratio()
                    if similarity >= 0.6:
                        return str(self.pcode_list[i]), "", similarity
        
        return "", "", 0.0


class TextProcessor:
    """Main text processing pipeline orchestrator"""
    
    def __init__(self, pcode_list: List[str], debug: bool = False):
        self.debug = debug
        self.cleanup = RawCleanupLayer(debug)
        self.normalizer = LinguisticNormalizer(debug)
        # Build allowed lengths from product list (remove dots)
        allowed_lengths = set(len(str(p).replace('.', '')) for p in pcode_list)
        self.extractor = PatternExtractor(debug, allowed_lengths)
        self.matcher = Matcher(pcode_list, debug)
    
    def process(self, text: str) -> Tuple[str, str, float]:
        """Process text through all stages and return (product_code, quantity, confidence)"""
        try:
            # Stage 1: Raw cleanup
            cleaned = ErrorHandler.safe_execute(
                lambda: self.cleanup.process(text),
                ProcessingStage.RAW_CLEANUP,
                default_return=text
            )
            
            # Stage 2: Linguistic normalization
            normalized = ErrorHandler.safe_execute(
                lambda: self.normalizer.process(cleaned),
                ProcessingStage.LINGUISTIC_NORMALIZATION,
                default_return=cleaned
            )
            
            # Stage 3: Pattern extraction
            candidates = ErrorHandler.safe_execute(
                lambda: self.extractor.extract_candidates(normalized, text),
                ProcessingStage.PATTERN_EXTRACTION,
                default_return=[]
            )
            
            if not candidates:
                if self.debug:
                    print(f"  [PROCESSOR] No candidates found")
                return "", "", 0.0
            
            # Stage 4: Confidence & Recovery
            pcode, remainder, confidence = ErrorHandler.safe_execute(
                lambda: self.matcher.find_best_match(candidates),
                ProcessingStage.CONFIDENCE_RECOVERY,
                default_return=("", "", 0.0)
            )
            
            # Extract quantity
            qty = self._extract_quantity(normalized, pcode, remainder)
            
            return pcode, qty, confidence
            
        except Exception as e:
            ErrorHandler.log_error(ProcessingStage.PATTERN_EXTRACTION, e, f"Processing: {text}")
            return "", "", 0.0
    
    def _extract_quantity(self, normalized: str, pcode_match: str, remainder: str) -> str:
        """Extract quantity from remainder or normalized text"""
        # If remainder is a valid quantity
        if remainder and remainder.isdigit():
            return remainder
        
        # Look for quantity after product code
        if not pcode_match:
            return ""
        
        pcode_clean = pcode_match.replace('.', '').lower()
        normalized_clean = normalized.replace(' ', '').lower()
        
        idx = normalized_clean.find(pcode_clean)
        if idx != -1:
            after_pcode = normalized_clean[idx + len(pcode_clean):]
            qty_match = re.search(r'(\d+)', after_pcode)
            if qty_match:
                return qty_match.group(1)
        
        return ""


class LiveTranscriber:
    """Live transcriber focused on audio handling and coordination"""
    
    def __init__(self, excel_file: str = "productlist.xlsx", language_code: str = "en-US", debug: bool = False):
        self.should_stop = False
        self.language_code = language_code
        self.debug = debug
        
        # Load product codes
        excel_path = os.path.join(os.path.dirname(__file__), excel_file)
        pcode_list = ErrorHandler.safe_execute(
            lambda: self._load_pcodes_from_excel(excel_path),
            ProcessingStage.RAW_CLEANUP,
            default_return=[]
        )
        
        if not pcode_list:
            print(f"❌ Warning: No product codes loaded!")
        
        print(f"✅ Loaded {len(pcode_list)} product codes")
        
        # Initialize processor
        self.processor = TextProcessor(pcode_list, debug)
        
        # Audio buffer
        self.audio_buffer = []
        self.last_audio_ts: float = time.time()
        self.is_recording = False
        
        # Google Cloud Speech client
        try:
            self.client = speech.SpeechClient()
        except Exception as e:
            ErrorHandler.log_error(ProcessingStage.RAW_CLEANUP, e, "Initializing Google Speech client")
            raise
    
    def _load_pcodes_from_excel(self, excel_file: str) -> List[str]:
        """Load product codes from Excel file"""
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            pcodes = []
            for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
                if row[0]:
                    code = str(row[0]).strip()
                    pcodes.append(code)
            
            wb.close()
            return pcodes
        except Exception as e:
            ErrorHandler.log_error(ProcessingStage.RAW_CLEANUP, e, f"Loading Excel: {excel_file}")
            return []
    
    def _silence_detector(self, audio_data: np.ndarray) -> bool:
        """Detect if audio is silence"""
        if audio_data.size == 0:
            return True
        rms = np.sqrt(np.mean(audio_data.astype(np.float32) ** 2))
        return rms < SILENCE_AMPLITUDE
    
    def _audio_callback(self, indata: np.ndarray, frames: int, _time, status) -> None:
        """Audio input callback"""
        if status:
            pass  # Suppress overflow messages
        
        now = time.time()
        self.audio_buffer.extend(indata.flatten().tolist())
        
        if self._silence_detector(indata):
            if self.is_recording and (now - self.last_audio_ts) >= SILENCE_SECONDS:
                self._process_audio_chunk()
                self.audio_buffer = []
                self.is_recording = False
        else:
            self.is_recording = True
            self.last_audio_ts = now
    
    def _process_audio_chunk(self) -> None:
        """Process audio chunk using Google Speech-to-Text"""
        if not self.audio_buffer:
            return
        
        audio_data = np.array(self.audio_buffer, dtype=np.int16)
        temp_filename = None
        
        try:
            # Write WAV file
            with tempfile.NamedTemporaryFile(suffix='.wav', delete=False) as temp_file:
                temp_filename = temp_file.name
            
            with wave.open(temp_filename, 'wb') as wav_file:
                wav_file.setnchannels(CHANNELS)
                wav_file.setsampwidth(2)
                wav_file.setframerate(SAMPLE_RATE)
                wav_file.writeframes(audio_data.tobytes())
            
            # Send to Google Cloud Speech API
            with open(temp_filename, 'rb') as f:
                content = f.read()
            
            audio = speech.RecognitionAudio(content=content)
            config = speech.RecognitionConfig(
                encoding=speech.RecognitionConfig.AudioEncoding.LINEAR16,
                sample_rate_hertz=SAMPLE_RATE,
                language_code=self.language_code,
                enable_automatic_punctuation=False,
                use_enhanced=True,
                model='default',
            )
            
            response = ErrorHandler.safe_execute(
                lambda: self.client.recognize(config=config, audio=audio),
                ProcessingStage.RAW_CLEANUP,
                default_return=None
            )
            
            if not response:
                return
            
            transcript_text = ""
            if response.results:
                transcript_text = " ".join(r.alternatives[0].transcript for r in response.results).strip()
            
            if transcript_text.strip():
                print(f"\n{'='*70}")
                print(f"🎤 LIVE: {transcript_text}")
                print(f"{'='*70}")
                
                # Process transcript
                pcode, qty, confidence = ErrorHandler.safe_execute(
                    lambda: self.processor.process(transcript_text),
                    ProcessingStage.CONFIDENCE_RECOVERY,
                    default_return=("", "", 0.0)
                )
                
                if pcode:
                    conf_str = f" (Confidence: {confidence:.2f})" if confidence < 1.0 else ""
                    print(f"\n{'─'*70}")
                    print(f"✅ PRODUCT: {pcode.upper()}  |  QTY: {qty or '-'}{conf_str}")
                    print(f"{'─'*70}\n")
                else:
                    print(f"\n{'─'*70}")
                    print(f"❌ PRODUCT: Not recognized")
                    print(f"{'─'*70}\n")
            else:
                print(f"\n🔇 [silence/no speech]\n")
        
        except Exception as e:
            ErrorHandler.log_error(ProcessingStage.CONFIDENCE_RECOVERY, e, "Processing audio chunk")
        finally:
            # Cleanup temp file
            if temp_filename:
                try:
                    os.unlink(temp_filename)
                except:
                    pass
    
    def run(self) -> None:
        """Run live transcription"""
        print("\n" + "="*70)
        print("🎤 LIVE TRANSCRIBER - Google Speech-to-Text")
        print("="*70)
        print("Press CTRL+C to stop\n")
        
        try:
            with sd.InputStream(
                samplerate=SAMPLE_RATE,
                channels=CHANNELS,
                dtype=DTYPE,
                callback=self._audio_callback,
                blocksize=int(SAMPLE_RATE * 0.1)
            ):
                while not self.should_stop:
                    time.sleep(0.1)
        
        except KeyboardInterrupt:
            print("\n" + "="*70)
            print("🛑 STOPPED")
            print("="*70 + "\n")
        except Exception as e:
            ErrorHandler.log_error(ProcessingStage.RAW_CLEANUP, e, "Audio stream error")


def test_mode(sample_texts: List[str], excel_file: str = "productlist.xlsx", debug: bool = True):
    """Test mode: process sample texts without audio"""
    print("\n" + "="*70)
    print("🧪 TEST MODE - Processing sample texts")
    print("="*70 + "\n")
    
    # Load product codes
    excel_path = os.path.join(os.path.dirname(__file__), excel_file)
    pcode_list = []
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
            if row[0]:
                pcode_list.append(str(row[0]).strip())
        wb.close()
    except Exception as e:
        print(f"❌ Error loading Excel: {e}")
        return
    
    processor = TextProcessor(pcode_list, debug=debug)
    
    for i, sample in enumerate(sample_texts, 1):
        print(f"\n{'─'*70}")
        print(f"Test {i}: '{sample}'")
        print(f"{'─'*70}")
        
        pcode, qty, confidence = processor.process(sample)
        
        conf_str = f" (Confidence: {confidence:.2f})" if confidence < 1.0 else ""
        if pcode:
            print(f"✅ Result: Product={pcode.upper()}, Quantity={qty or '-'}{conf_str}")
        else:
            print(f"❌ Result: Not recognized (Confidence: {confidence:.2f})")


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(description="Live Speech-to-Text Transcriber")
    parser.add_argument("--test", action="store_true", help="Run in test mode with sample texts")
    parser.add_argument("--debug", action="store_true", help="Enable debug output")
    parser.add_argument("--text", type=str, help="Test with specific text")
    args = parser.parse_args()
    
    global DEBUG
    DEBUG = args.debug
    
    if args.test or args.text:
        # Test mode
        if args.text:
            test_texts = [args.text]
        else:
            # Default test cases
            test_texts = [
                "double 2 3",          # Should become: 223
                "s dot ss dot de",     # Should become: s.ss.de
                "in c 35 335",         # Should parse as: INC35335 (keep "in" as "in")
                "z triple 0 3",        # Should become: z0003 (not z03)
                "n c 35 335",          # Should parse as: NC35335 (keep "n" as "n")
                "z0003",               # Should stay: z0003
                # NEW: Test cases based on actual product codes
                "a 0 0 0 2",           # Should become: A0002 (single letter pattern)
                "i p 0 0 0 0 1",      # Should become: IP00001 (multi-letter pattern)
                "j c dot n p p",       # Should become: JC.NPP (dot pattern)
                "double zero zero zero one", # Should become: 0001 (pure numeric)
                "h 0 2 8 2",           # Should become: H0282 (single letter)
                "i n b a c a 0 1",     # Should become: INBACA01 (multi-letter)
            ]
        
        test_mode(test_texts, debug=DEBUG)
    else:
        # Live mode
        transcriber = LiveTranscriber(debug=DEBUG)
        transcriber.run()


if __name__ == "__main__":
    main()
